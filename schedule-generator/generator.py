import openpyxl
import firebase_admin
from firebase_admin import credentials, firestore
from datetime import datetime, timezone, timedelta


def delete_collection(coll_ref):
    docs = coll_ref.list_documents()
    deleted = 0

    for doc in docs:
        doc.delete()
        deleted = deleted + 1


def createGame(db, gameName, gameTime, field, _type, WNextGame, LNextGame, team1Name, team2Name):
    gameObject = {
        'GameName': gameName,
        'Team1Score': 0,
        'Team2Score': 0,
        'Status': 0,
        'WNextGame': WNextGame,
        'LNextGame': LNextGame,
        'DateTime': gameTime,
        'Field': field,
        'Ready': 1,
        'Type': _type,
        'Team1Name': team1Name,
        'Team2Name': team2Name,
    }
    doc_ref = db.collection('Games').add(gameObject)

    return doc_ref[1].id

def generateDateTime(year, month, day, hours, minutes):
    dateTime = datetime(year, month, day, hours, minutes)
    # Define a fixed offset timezone (e.g., UTC+2)
    fixed_offset = timezone(timedelta(hours=2))
    # Assign the fixed offset timezone to the naive datetime object
    aware_datetime = dateTime.replace(tzinfo=fixed_offset)
    return aware_datetime

def generate_bracket_games(db, wb, date):
    ws = wb['Slutspel']
    
    year = int(date.year)
    month = int(date.month)
    day = int(date.day)

    #FINAL
    time = ws['K13'].value
    hours = int(time.hour)
    minutes = int(time.minute)
    name = ws['L13'].value
    field = ws['M13'].value 
    _type = 1
    WNextGame = []
    LNextGame = []
    team1Name = ws['L15'].value 
    team2Name = ws['M15'].value

    dateTime = generateDateTime(year, month, day, hours, minutes)
    
    finalIndex = createGame(db, name, dateTime, field, _type, WNextGame, LNextGame, team1Name, team2Name)
    
    
    #BRONS
    time = ws['K9'].value
    hours = int(time.hour)
    minutes = int(time.minute)
    name = ws['L9'].value
    field = ws['M9'].value 
    _type = 1
    WNextGame = []
    LNextGame = []
    team1Name = ws['L11'].value 
    team2Name = ws['M11'].value

    dateTime = generateDateTime(year, month, day, hours, minutes)
    
    bronzeIndex = createGame(db, name, dateTime, field, _type, WNextGame, LNextGame, team1Name, team2Name)
    

    #SEMI 1
    time = ws['F9'].value
    hours = int(time.hour)
    minutes = int(time.minute)
    name = ws['G9'].value
    field = ws['H9'].value 
    _type = 1
    WNextGame = [finalIndex, 1]
    LNextGame = [bronzeIndex, 1]
    team1Name = ws['G11'].value 
    team2Name = ws['H11'].value

    dateTime = generateDateTime(year, month, day, hours, minutes)
    
    semi1Index = createGame(db, name, dateTime, field, _type, WNextGame, LNextGame, team1Name, team2Name)

    #SEMI 2
    time = ws['F13'].value
    hours = int(time.hour)
    minutes = int(time.minute)
    name = ws['G13'].value
    field = ws['H13'].value 
    _type = 1
    WNextGame = [finalIndex, 2]
    LNextGame = [bronzeIndex, 2]
    team1Name = ws['G15'].value 
    team2Name = ws['H15'].value

    dateTime = generateDateTime(year, month, day, hours, minutes)
    
    semi2Index = createGame(db, name, dateTime, field, _type, WNextGame, LNextGame, team1Name, team2Name)

    
    
    #KVART 1
    time = ws['A5'].value
    hours = int(time.hour)
    minutes = int(time.minute)
    name = ws['B5'].value
    field = ws['C5'].value 
    _type = 1
    WNextGame = [semi1Index, 1]
    LNextGame = []
    team1Name = ws['B7'].value 
    team2Name = ws['C7'].value

    dateTime = generateDateTime(year, month, day, hours, minutes)
    
    qf1Index = createGame(db, name, dateTime, field, _type, WNextGame, LNextGame, team1Name, team2Name)
        
    #KVART 2
    time = ws['A9'].value
    hours = int(time.hour)
    minutes = int(time.minute)
    name = ws['B9'].value
    field = ws['C9'].value 
    _type = 1
    WNextGame = [semi2Index, 1]
    LNextGame = []
    team1Name = ws['B11'].value 
    team2Name = ws['C11'].value

    dateTime = generateDateTime(year, month, day, hours, minutes)
    
    qf2Index = createGame(db, name, dateTime, field, _type, WNextGame, LNextGame, team1Name, team2Name)

    #KVART 3
    time = ws['A13'].value
    hours = int(time.hour)
    minutes = int(time.minute)
    name = ws['B13'].value
    field = ws['C13'].value 
    _type = 1
    WNextGame = [semi1Index, 2]
    LNextGame = []
    team1Name = ws['B15'].value 
    team2Name = ws['C15'].value

    dateTime = generateDateTime(year, month, day, hours, minutes)
    
    qf3Index = createGame(db, name, dateTime, field, _type, WNextGame, LNextGame, team1Name, team2Name)

    #KVART 4
    time = ws['A17'].value
    hours = int(time.hour)
    minutes = int(time.minute)
    name = ws['B17'].value
    field = ws['C17'].value 
    _type = 1
    WNextGame = [semi2Index, 2]
    LNextGame = []
    team1Name = ws['B19'].value 
    team2Name = ws['C19'].value

    dateTime = generateDateTime(year, month, day, hours, minutes)
    
    qf4Index = createGame(db, name, dateTime, field, _type, WNextGame, LNextGame, team1Name, team2Name)
         

    # ID FÖR KVARTSFINALER [KVART1, ... , KVART4]
    return [qf1Index, qf2Index, qf3Index, qf4Index]

def parseTeamsInGroup(wb, col, row):
    ws = wb['Gruppspel']
    teams = []

    moreTeams = True
    while(moreTeams):
        cell = col + str(row)
        if(ws[cell].value != None):
            teams.append(ws[cell].value)
        else:
            moreTeams = False    
        row += 1
    return teams

def createTeams(db, teamList):
    teamIDmap = {}

    #Skapa lag (Spara map namn till ID)
    for t in teamList:
        team = {
            'TeamName': t
        }
        doc_ref = db.collection('Teams').add(team)
        teamIDmap[t] = doc_ref[1].id
    
    return teamIDmap

def createGroup(db, groupName, nextGames):
    group = {
        'GroupName': groupName,
        'NextGames': nextGames,
        'GameIDs': [],
    }

    doc_ref = db.collection('Groups').add(group)
    return doc_ref[1].id 

def createGroupTeams(db, teams, teamIDmap, groupID, groupName):
    count = 1
    for t in teams:
        teamData = [t, 0, 0, 0, 0, 0, count]

        groupTeam = {
            'TeamID': teamIDmap[t],
            'GroupID': groupID,
            'TeamData': teamData,
            'GroupName': groupName,
        }

        count += 1
        db.collection('GroupTeams').add(groupTeam)

def createTeamGame(db, teamID, gameID, position):
    teamGame = {
        'TeamID': teamID,
        'GameID': gameID,
        'TeamPosition': position,
    }

    db.collection('TeamGame').add(teamGame)

def generateGroupGames(db, wb, col, gameRow, groupName, date, teamIDmap):
    ws = wb['Gruppspel']
    col = ord(col)
    gameIDs = []

    moreGames = True
    nameCounter = 1
    while(moreGames):
        team1name = ws[chr(col)+str(gameRow)].value
        team2name = ws[chr(col+1)+str(gameRow)].value
        time = ws[chr(col-1)+str(gameRow-2)].value
        dateTime = generateDateTime(int(date.year), int(date.month), int(date.day), int(time.hour), int(time.minute))
        field = ws[chr(col+2)+str(gameRow-2)].value
        name = groupName + str(nameCounter)
        nameCounter += 1
        gameRow += 4

        #Skapa match
        gameID = createGame(db, name, dateTime, field, 0, [], [], team1name, team2name)

        #Skapa teamgame
        createTeamGame(db, teamIDmap[team1name], gameID, 1)
        createTeamGame(db, teamIDmap[team2name], gameID, 2)

        #Lägg till id i lista
        gameIDs.append(gameID)

        if(ws[chr(col)+str(gameRow)].value == None):
            moreGames = False
    
    return gameIDs
    
def generateGroup(db, wb, col, teamRow, gameRow, groupName, nextGames, date):
    #Läs in lag
    teams = parseTeamsInGroup(wb, col, teamRow)

    #Skapa lag
    teamIDmap = createTeams(db, teams)

    #Skapa gruppobjekt
    groupID = createGroup(db, groupName, nextGames)

    #Lägg till lag i gruppen
    createGroupTeams(db, teams, teamIDmap, groupID, groupName)

    #Skapa match
    #Skapa teamGame
    gameIDs = generateGroupGames(db, wb, col, gameRow, groupName, date, teamIDmap)

    #Lägg till gameids i group
    groupRef = db.collection("Groups").document(groupID)
    groupRef.update({'GameIDs': gameIDs})

def generateAllGroups(db, wb, date, qf):
    groupNames = ['A', 'B', 'C', 'D']
    columns = ['B', 'F', 'J', 'N']
    teamRow = 4
    gameRow = 13
    nextGames = [[qf[0], 1, qf[1], 1], 
                 [qf[1], 2, qf[0], 2],
                 [qf[2], 1, qf[3], 1],
                 [qf[3], 2, qf[2], 2]]

    for i in range(len(groupNames)):
        generateGroup(db, wb, columns[i], teamRow, gameRow, groupNames[i], nextGames[i], date)

def deleteDBDocuments(db):
    gameRef = db.collection('Games')
    teamRef = db.collection('Teams')
    groupRef = db.collection('Groups')
    gtRef = db.collection('GroupTeams')
    tgRef = db.collection('TeamGame')

    delete_collection(gameRef)
    delete_collection(teamRef)
    delete_collection(groupRef)
    delete_collection(gtRef)
    delete_collection(tgRef)
    
cred_path = './service_key.json'

# Use the application default credentials
cred = credentials.Certificate(cred_path)

# Initialize the Firestore client
firebase_admin.initialize_app(cred)
db = firestore.client()

deleteDBDocuments(db)

wb = openpyxl.load_workbook('schedule.xlsx', data_only=True)

ws2 = wb['Gruppspel']
date = ws2['A1'].value

qf = generate_bracket_games(db, wb, date)
generateAllGroups(db, wb, date, qf)
