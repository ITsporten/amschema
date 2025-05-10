import openpyxl
import firebase_admin
from firebase_admin import credentials, firestore

def delete_collection(coll_ref):
    docs = coll_ref.list_documents()
    deleted = 0

    for doc in docs:
        doc.delete()
        deleted = deleted + 1

def deleteDBDocuments(db):
    standingRef = db.collection('Standings')
    delete_collection(standingRef)

def createStandingItem(db, ws, col, row, seasonName):
    name = ws[col+str(row)].value
    name = name.strip()
    
    col = ord(col)

    points = []
    totalPoints = 0

    for i in range(1, 5):
        value = ws[chr(col+i) + str(row)].value
        if(value == None):
            value = 0
        
        points.append(value)
        totalPoints += value
    
    standingItem = {
        'TeamName': name,
        'Points': points,
        'TotalPoints': totalPoints,
        'Season': seasonName
    }
    db.collection('Standings').add(standingItem)

def createAllStandings(db, ws, col, row, seasonName):
    while(True):
        if(ws[col+str(row)].value == None):
            break

        createStandingItem(db, ws, col, row, seasonName)
        row += 1


if __name__ == "__main__":
    cred_path = './service_key.json'

    # Use the application default credentials
    cred = credentials.Certificate(cred_path)

    # Initialize the Firestore client
    firebase_admin.initialize_app(cred)
    db = firestore.client()

    wb = openpyxl.load_workbook('./datasheets/standings.xlsx', data_only=True)
    ws = wb.active

    deleteDBDocuments(db)

    seasonName = (ws['B4'].value).strip()
    createAllStandings(db, ws, 'D', 4, seasonName)