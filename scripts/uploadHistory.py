import openpyxl
import firebase_admin
from firebase_admin import credentials, firestore

def delete_collection(coll_ref):
    docs = coll_ref.list_documents()
    deleted = 0

    for doc in docs:
        doc.delete()
        deleted = deleted + 1

def deleteDBDocuments(db):
    historyRef = db.collection('History')
    delete_collection(historyRef)

def createHistoryItem(db, ws, col, row):
    season = ws[col+str(row)].value
    col = ord(col)

    points = []
    standings = []
    placements = []

    standingCol = chr(col+1)
    pointCol = chr(col+2)
    
    for i in range(3):
        row += 1
        teamName = ws[standingCol+str(row)].value
        if(teamName == '-'):
            teamName = "?"
        standings.append(teamName)

        point = ws[pointCol+str(row)].value
        if(point != '-'):
            point = int(point)
        points.append(point)

        placement = int(ws[chr(col)+str(row)].value)
        placements.append(placement)

    historyItem = {
        'Season': season,
        'TeamName': standings,
        'Points': points,
        'Placements': placements,
    }

    db.collection('History').add(historyItem)

def createAllHistory(db, ws, col, row):
    while(True):
        if ws[col+str(row)].value == None:
            break

        createHistoryItem(db, ws, col, row)
        row += 5

if __name__ == "__main__":
    cred_path = './service_key.json'

    # Use the application default credentials
    cred = credentials.Certificate(cred_path)

    # Initialize the Firestore client
    firebase_admin.initialize_app(cred)
    db = firestore.client()

    wb = openpyxl.load_workbook('./datasheets/history.xlsx', data_only=True)
    ws = wb.active

    deleteDBDocuments(db)

    createAllHistory(db, ws, 'A', 1)